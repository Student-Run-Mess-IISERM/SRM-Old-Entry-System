VERSION="1.3.4"

import os
import json
import traceback
from datetime import datetime, timedelta

import gspread
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
from oauth2client.service_account import ServiceAccountCredentials
from customtkinter import (
    CTk,
    CTkButton,
    CTkCheckBox,
    CTkEntry,
    CTkFrame,
    CTkFont,
    CTkLabel,
    CTkSegmentedButton,
    CTkTabview,
    CTkTextbox,
    set_appearance_mode,
    set_window_scaling,
)
from tkinter import IntVar
from urllib3.exceptions import NewConnectionError

MealType = str
StatusLevel = str
Spreadsheet = gspread.spreadsheet.Spreadsheet
Worksheet = gspread.worksheet.Worksheet

now = datetime.now
strptime = datetime.strptime

GOOGLE_API_SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/drive',
]

try:
    gsheet_credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', GOOGLE_API_SCOPES)
    gsheet_client = gspread.authorize(gsheet_credentials)
    ONLINE_ENABLED = True
except FileNotFoundError:
    gsheet_client = None
    ONLINE_ENABLED = False
    print("credentials.json not found.  Online features disabled.")

NEXT_DAY_DATE = now() + timedelta(days=1)
NEXT_DAY_DATE_STRING = NEXT_DAY_DATE.strftime("%d %B, %Y")

CURRENT_DATE = now()
CURRENT_DATE_STRING = CURRENT_DATE.strftime("%d %B, %Y")

MEAL_COLUMN_MAPPING = {
    'Breakfast': {
        'status': 2,
        'time': 3,
    },
    'Lunch': {
        'status': 4,
        'time': 5,
    },
    'Dinner': {
        'status': 6,
        'time': 7,
    },
}

def column_values(worksheet, column):
    if isinstance(worksheet, Worksheet):
        range_label = f"{chr(64 + column)}:{chr(64 + column)}"
        return [row[0] if row else '' for row in worksheet.get(range_label)]
    elif isinstance(worksheet, xl.worksheet.worksheet.Worksheet):
        value_generator = worksheet.iter_cols(min_col=column, max_col=column, min_row=2, values_only=True)
        return [value for value in next(value_generator, [])]
    else:
        raise TypeError("Unsupported worksheet type")

def row_values(worksheet, row):
    if isinstance(worksheet, Worksheet):
        return worksheet.row_values(row)
    elif isinstance(worksheet, xl.worksheet.worksheet.Worksheet):
        value_generator = worksheet.iter_rows(min_row=row, max_row=row, values_only=True)
        return list(next(value_generator, []))
    else:
        raise TypeError("Unsupported worksheet type")

def gsheet_batch_upload(sheet, header, data):
    if not ONLINE_ENABLED:
        return
    sheet.clear()
    sheet.append_row(header)
    if not data:
        return
    shape_data = (len(data), len(data[0]))
    max_row_number = shape_data[0] + 1
    max_col_letter = chr(65 + shape_data[1] - 1)
    range_name = f'A2:{max_col_letter}{max_row_number}'
    sheet.update(range_name=range_name, values=data)

def leave_update():
    leave_details_spreadsheet = gsheet_client.open('Leave Details for SRM')
    current_leave_details_worksheet = leave_details_spreadsheet.worksheet('Current Leave Details')
    all_leaves_worksheet = leave_details_spreadsheet.worksheet('Form Responses 1')
    current_leave_details_worksheet.clear()
    all_leave_values = all_leaves_worksheet.get_all_values()
    leave_list_header = all_leave_values[0]

    leave_data = []
    for leave_detail in all_leave_values[1:]:
        try:
            start_date = strptime(leave_detail[5], '%m/%d/%Y')
            end_date = strptime(leave_detail[6], '%m/%d/%Y')
        except ValueError:
            continue
        is_today_leave = (NEXT_DAY_DATE - start_date).days >= 0 and (end_date - NEXT_DAY_DATE).days >= 0
        if is_today_leave:
            leave_data.append(leave_detail)

    if len(leave_data) == 0:
        return

    gsheet_batch_upload(current_leave_details_worksheet, leave_list_header, leave_data)

class Repository:
    def __init__(self, repository_worksheet):
        if not ONLINE_ENABLED:
            return
        values_column = [cell_value.split(",") for cell_value in repository_worksheet.col_values(2)]
        self.file_names = values_column[0]
        self.sheet_names = values_column[1]
        self.name_columns = [int(column) for column in values_column[2]]
        self.registration_number_columns = [int(column) for column in values_column[3]]
        self.meal_columns = [int(column) for column in values_column[4]]
        self.share_to_emails = [email.strip() for email in values_column[5] if email.strip() != '']

def subscriber_data_update():
    if not ONLINE_ENABLED:
        return
    repository_details_worksheet = gsheet_client.open('Repository Details for SRM').worksheet('Sheet1')
    repository = Repository(repository_details_worksheet)
    subscriber_repository_worksheet = gsheet_client.open('Repository for SRM').worksheet('Sheet1')
    subscriber_repository_worksheet.clear()
    subscriber_repository_header = ['Student Name', 'Registration Number', 'Meals Opted']
    all_subscribers = []

    for file, sheet in zip(repository.file_names, repository.sheet_names):
        subscriber_worksheet = gsheet_client.open(file).worksheet(sheet)
        subscribers = subscriber_worksheet.get_all_values()

        for subscriber_detail in subscribers[1:]:
            all_subscribers.append([
                subscriber_detail[repository.name_columns[0]],
                subscriber_detail[repository.registration_number_columns[0]].split('@')[0],
                subscriber_detail[repository.meal_columns[0]],
            ])

    gsheet_batch_upload(subscriber_repository_worksheet, subscriber_repository_header, all_subscribers)

    if not os.path.exists('Subscriber Data.xlsx'):
        subscriber_workbook = xl.Workbook()
        subscriber_workbook.remove(subscriber_workbook['Sheet'])
        subscriber_workbook.create_sheet('Subscriber Data')
        subscriber_sheet = subscriber_workbook['Subscriber Data']
    else:
        subscriber_workbook = xl.load_workbook('Subscriber Data.xlsx')
        subscriber_workbook.remove(subscriber_workbook['Subscriber Data'])
        subscriber_workbook.create_sheet('Subscriber Data')
        subscriber_sheet = subscriber_workbook['Subscriber Data']

    subscriber_sheet_header = [
        'Student Name', 'Registration Number',
        'Breakfast', 'Lunch', 'Dinner'
    ]
    subscriber_sheet.append(subscriber_sheet_header)

    for subscriber_data in enumerate(all_subscribers, start=2):
        row = [
            subscriber_data[1][0],
            subscriber_data[1][1].upper().strip(),
            'NOT' if 'Breakfast' not in subscriber_data[1][2].split(', ') else '',
            'NOT' if 'Lunch' not in subscriber_data[1][2].split(', ') else '',
            'NOT' if 'Dinner' not in subscriber_data[1][2].split(', ') else '',
        ]
        subscriber_sheet.append(row)

    subscriber_workbook.save('Subscriber Data.xlsx')

class App(CTk):
    def __init__(self):
        super().__init__()
        self.meal_colors = self.generate_meal_colors()
        self.title(f'SRM Data Entry System v{VERSION}')
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        self.sidebar_frame = CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)

        self.logo_label = CTkLabel(
            self.sidebar_frame,
            text="Student Run Mess",
            font=CTkFont(size=20, weight="bold")
        )
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        set_appearance_mode("Dark")

        self.status = CTkEntry(self, placeholder_text="Status")
        self.status.configure(state='readonly')
        self.status.grid(row=3, column=1, columnspan=3, padx=(20, 20), pady=(20, 20), sticky="nsew")

        self.tabview = CTkTabview(self)
        self.tabview.grid(row=0, rowspan=3, column=1, padx=(20, 20), pady=(20, 0), sticky="nsew")

        self.tabview.add("Daily Entry")
        self.tabview.tab("Daily Entry").grid_columnconfigure((0, 1), weight=1)
        
        self.prepaid_entry = CTkFrame(self.tabview.tab("Daily Entry"))
        self.prepaid_entry.grid(row=0, column=0, columnspan=3, padx=(20, 10), pady=(20, 10), sticky="nsew")
        self.prepaid_entry.grid_rowconfigure((0, 1, 2, 3), weight=1)
        self.prepaid_entry.grid_columnconfigure(0, weight=1)
        self.prepaid_entry.grid_columnconfigure(1, weight=4)
        self.prepaid_entry.grid_columnconfigure(2, weight=1)
        CTkLabel(self.prepaid_entry, text='MS25').grid(row=0, column=0, padx=(20, 0), pady=(20, 10), sticky="nsew")
        CTkLabel(self.prepaid_entry, text='MS24').grid(row=1, column=0, padx=(20, 0), pady=(20, 10), sticky="nsew")
        CTkLabel(self.prepaid_entry, text='MS23').grid(row=2, column=0, padx=(20, 0), pady=(10, 10), sticky="nsew")
        CTkLabel(self.prepaid_entry, text='Others').grid(row=3, column=0, padx=(20, 0), pady=(10, 20), sticky="nsew")
        self.ms25 = CTkEntry(self.prepaid_entry, width=200)
        self.ms25.grid(row=0, column=1, padx=(20, 0), pady=(20, 10), sticky="nsew")
        self.ms24 = CTkEntry(self.prepaid_entry, width=200)
        self.ms24.grid(row=1, column=1, padx=(20, 0), pady=(20, 10), sticky="nsew")
        self.ms23 = CTkEntry(self.prepaid_entry)
        self.ms23.grid(row=2, column=1, padx=(20, 0), pady=(10, 10), sticky="nsew")
        self.others = CTkEntry(self.prepaid_entry)
        self.others.grid(row=3, column=1, padx=(20, 0), pady=(10, 20), sticky="nsew")

        self.coupon_entry = CTkFrame(self.tabview.tab("Daily Entry"))
        self.coupon_entry.grid(row=0, column=3, columnspan=3, padx=(10, 20), pady=(20, 10), sticky="nsew")
        self.coupon_entry.grid_columnconfigure(0, weight=1)
        self.coupon_entry.grid_columnconfigure(1, weight=4)
        CTkLabel(self.coupon_entry, text='Coupon').grid(row=0, column=0, padx=(20, 0), pady=(20, 10), sticky="nsw")
        CTkLabel(self.coupon_entry, text='Amount').grid(row=1, column=0, padx=(20, 0), pady=(10, 10), sticky="nsw")
        CTkLabel(self.coupon_entry, text='Coupons Sold').grid(row=2, column=0, padx=(20, 0), pady=(10, 20), sticky="nsw")

        self.coupon = CTkEntry(self.coupon_entry)
        self.coupon.grid(row=0, column=1, columnspan=2, padx=(20, 20), pady=(20, 10), sticky="nsew")
        self.amount = CTkEntry(self.coupon_entry)
        self.amount.grid(row=1, column=1, columnspan=2, padx=(20, 20), pady=(10, 10), sticky="nsew")
        self.coupons_sold = CTkEntry(self.coupon_entry)
        self.coupons_sold.grid(row=2, column=1, padx=(20, 0), pady=(10, 20), sticky="nsew")
        self.coupons_sold.insert(0, '0')
        self.coupons_sold.configure(state='readonly')

        self.extra_config = CTkFrame(self.tabview.tab("Daily Entry"))
        self.extra_config.grid(row=1, column=4, columnspan=2, padx=(10, 20), pady=(10, 20), sticky="nsew")
        self.extra_config.grid_columnconfigure(0, weight=1)
        self.update = IntVar(value=1)

        CTkCheckBox(self.extra_config, text='Update in Database', variable=self.update).grid(
            row=0, column=0, padx=(20, 20), pady=(20, 10), sticky='nsew'
        )
        
        self.details_frame = None

        self.config_frame = CTkFrame(self.tabview.tab("Daily Entry"))
        self.config_frame.grid(row=1, column=0, columnspan=4, padx=(20, 10), pady=(10, 20), sticky="nsew")
        self.config_frame.grid_columnconfigure(0, weight=1)

        self.details_frame = self.create_details_frame(self.tabview.tab("Daily Entry"))
        
        self.non_veg = IntVar()
        CTkCheckBox(self.config_frame, text='Non-Veg', variable=self.non_veg).grid(
            row=0, column=0, padx=(20, 20), pady=(20, 10), sticky='nsew'
        )

        CTkLabel(self.config_frame, text='Extra price for Prepaid for Non-veg').grid(
            row=1, column=0, padx=(20, 20), pady=(10, 10), sticky='nsw'
        )
        self.prepaid_extra_price = CTkEntry(self.config_frame)
        self.prepaid_extra_price.grid(row=1, column=1, padx=(0, 20), pady=(10, 10), sticky="nsew")
        self.prepaid_extra_price.insert(0, '30')

        self.meal = CTkSegmentedButton(self.config_frame)
        self.meal.grid(row=3, column=0, columnspan=2, padx=(20, 20), pady=(10, 20), sticky='nsew')
        self.meal.configure(values=['Breakfast', 'Lunch', 'Dinner'])

        current_hour = now().hour
        if current_hour < 11:
            self.meal.set('Breakfast')
        elif 11 <= current_hour <= 17:
            self.meal.set('Lunch')
        else:
            self.meal.set('Dinner')

        try:
            with open('constants.json', 'r') as f:
                self.constants = json.load(f)
        except FileNotFoundError:
            self.constants = {
                'hostel_number': 5,
            }

        self.hostel = CTkLabel(
            self.config_frame,
            text=f"Hostel {self.constants['hostel_number']}",
        )
        self.hostel.grid(row=4, column=0, columnspan=2, padx=(20, 20), pady=(10, 10), sticky='nsew')
        
        scaling_factor = self.constants.get("scaling", 0.8)
        if not isinstance(scaling_factor, (int, float)):
            self.write_to_status_bar("Error: 'scaling' value in constants.json is not a number. Using default scaling (1).", level='error')
            scaling_factor = 1
        elif not (0.5 <= scaling_factor <= 3.0):
            self.write_to_status_bar("Warning: 'scaling' value in constants.json is outside the recommended range (0.5 to 2.0).  Using the provided value.", level='warning')
        set_window_scaling(scaling_factor)

        self.tabview.add("Create File")
        self.tabview.tab("Create File").grid_columnconfigure((0, 1), weight=1)
        self.create_file = CTkFrame(self.tabview.tab("Create File"))
        self.create_file.grid(row=0, column=0, padx=(20, 10), pady=(20, 20), sticky="nsew")
        self.create_file.grid_columnconfigure((0, 1), weight=1)
        self.create_file.grid_rowconfigure((0, 1, 2), weight=1)
        CTkLabel(self.create_file, text="File Name").grid(
            row=0, column=0, padx=(20, 10), pady=(20, 10), sticky="nsw"
        )
        CTkLabel(self.create_file, text="Date").grid(
            row=1, column=0, padx=(20, 10), pady=(10, 10), sticky="nsw"
        )
        self.file_name = CTkEntry(self.create_file)
        self.file_name.grid(row=0, column=1, padx=(10, 20), pady=(20, 10), sticky="nse")
        self.file_name.insert(0, 'SRM Data')
        self.date = CTkEntry(self.create_file)
        self.date.grid(row=1, column=1, padx=(10, 20), pady=(10, 10), sticky="nse")

        self.create_database = IntVar(value=1)
        if current_hour >= 22:
            self.date.insert(0, NEXT_DAY_DATE_STRING)
        else:
            self.date.insert(0, CURRENT_DATE_STRING)

        self.update_leave_repository = IntVar(value=1)
        self.update_repository = IntVar(value=1)

        self.spreadsheet = CTkCheckBox(self.create_file, text='Google Spreadsheet', variable=self.create_database)
        self.spreadsheet.grid(row=3, column=0, columnspan=2, padx=(20, 10), pady=(10, 10), sticky="nsw")
        self.update_leave = CTkCheckBox(self.create_file, text='Update Leaves', variable=self.update_leave_repository)
        self.update_leave.grid(row=4, column=0, columnspan=2, padx=(20, 10), pady=(10, 10), sticky="nsw")
        self.update_rep = CTkCheckBox(self.create_file, text='Update Repositories', variable=self.update_repository)
        self.update_rep.grid(row=5, column=0, columnspan=2, padx=(20, 10), pady=(10, 10), sticky="nsw")
        
        self.calculate_button = CTkButton(self.create_file, text="Calculate", command=self.calculate)
        self.calculate_button.grid(row=6, column=0, columnspan=2, padx=(20, 10), pady=(10, 20), sticky="nsew")

        self.information_box = CTkTextbox(self.tabview.tab("Create File"), height=50)
        self.information_box.grid(row=0, column=1, padx=(10, 20), pady=(20, 20), sticky="nsew")
        self.information_box.configure(state='disabled')

        self.create_prepaid_entry = self.logger_create(self.create_prepaid_entry)
        self.generate_coupon = self.logger_create(self.generate_coupon)
        self.create_daily_file = self.logger_create(self.create_daily_file)
        
        self.on_click_add_ms25 = lambda: self.create_prepaid_entry("MS25")
        self.on_click_add_ms24 = lambda: self.create_prepaid_entry("MS24")
        self.on_click_add_ms23 = lambda: self.create_prepaid_entry("MS23")
        self.on_click_add_others = lambda: self.create_prepaid_entry("others")
        self.on_click_generate_for_button = lambda: self.generate_coupon(self.coupon.get(), self.amount.get())
        
        self.add_ms25 = CTkButton(self.prepaid_entry, text='Add', command=self.on_click_add_ms25, width=100)
        self.add_ms25.grid(row=0, column=2, padx=(20, 20), pady=(20, 10), sticky="nse")
        self.add_ms24 = CTkButton(self.prepaid_entry, text='Add', command=self.on_click_add_ms24, width=100)
        self.add_ms24.grid(row=1, column=2, padx=(20, 20), pady=(20, 10), sticky="nse")
        self.add_ms23 = CTkButton(self.prepaid_entry, text='Add', command=self.on_click_add_ms23, width=100)
        self.add_ms23.grid(row=2, column=2, padx=(20, 20), pady=(10, 10), sticky="nse")
        self.add_others = CTkButton(self.prepaid_entry, text='Add', command=self.on_click_add_others, width=100)
        self.add_others.grid(row=3, column=2, padx=(20, 20), pady=(10, 20), sticky="nse")
        self.generate = CTkButton(
            self.coupon_entry,
            text='Generate',
            command=self.on_click_generate_for_button,
            width=100
        )
        self.generate.grid(row=2, column=2, padx=(10, 20), pady=(10, 20), sticky="nse")
        self.create = CTkButton(self.create_file, text='Create', command=self.create_daily_file, width=100)
        self.create.grid(row=2, column=1, padx=(10, 20), pady=(10, 10), sticky="nse")

        self.file_name.bind('<Down>', lambda event: self.date.focus_set())
        self.date.bind('<Up>', lambda event: self.file_name.focus_set())
        self.ms25.bind('<Down>', lambda event: self.ms24.focus_set())
        self.ms25.bind('<Right>', lambda event: self.coupon.focus_set())
        self.ms24.bind('<Up>', lambda event: self.ms25.focus_set())
        self.ms24.bind('<Down>', lambda event: self.ms23.focus_set())
        self.ms24.bind('<Right>', lambda event: self.amount.focus_set())
        self.ms23.bind('<Right>', lambda event: self.amount.focus_set())
        self.ms23.bind('<Up>', lambda event: self.ms24.focus_set())
        self.ms23.bind('<Down>', lambda event: self.others.focus_set())
        self.others.bind('<Up>', lambda event: self.ms23.focus_set())
        self.others.bind('<Right>', lambda event: self.amount.focus_set())
        self.coupon.bind('<Down>', lambda event: self.amount.focus_set())
        self.coupon.bind('<Left>', lambda event: self.ms24.focus_set())
        self.amount.bind('<Left>', lambda event: self.ms23.focus_set())
        self.amount.bind('<Up>', lambda event: self.coupon.focus_set())
        self.amount.bind('<Down>', lambda event: self.others.focus_set())

        self.file_name.bind('<Return>', lambda event: self.date.focus_set())
        self.date.bind('<Return>', lambda event: self.create_daily_file())
        self.coupon.bind('<Return>', lambda event: self.amount.focus_set())
        self.amount.bind('<Return>', lambda event: self.generate_coupon(self.coupon.get(), self.amount.get()))
        self.ms25.bind('<Return>', lambda event: self.on_click_add_ms25())
        self.ms24.bind('<Return>', lambda event: self.on_click_add_ms24())
        self.ms23.bind('<Return>', lambda event: self.on_click_add_ms23())
        self.others.bind('<Return>', lambda event: self.on_click_add_others())
        
        self._workbook_cache = None
        self._gsheet_cache = None
            
    def logger_create(self, fun):
        def wrapper(*args, **kwargs):
            try:
                fun(*args, **kwargs)
            except Exception as e:
                to_write = f'Error: {e} \n {traceback.format_exc()}'
                self.write_to_status_bar(to_write, 'error')
        return wrapper
    
    def create_details_frame(self, parent):
        frame = CTkFrame(parent)
        frame.grid(row=2, column=0, columnspan=6, padx=(20, 20), pady=(10, 20), sticky="nsew")

        name_label = CTkLabel(frame, text="Name: ", text_color="black")
        name_label.grid(row=0, column=0, padx=(10, 5), pady=(5, 5), sticky="w")
        self.name_value = CTkLabel(frame, text="", text_color="black")
        self.name_value.grid(row=0, column=1, padx=(5, 10), pady=(5, 5), sticky="w")
        reg_label = CTkLabel(frame, text="Registration Number: ", text_color="black")
        reg_label.grid(row=1, column=0, padx=(10, 5), pady=(5, 5), sticky="w")
        self.reg_value = CTkLabel(frame, text="", text_color="black")
        self.reg_value.grid(row=1, column=1, padx=(5, 10), pady=(5, 5), sticky="w")

        meals_label = CTkLabel(frame, text="Number of Meals Subscribed To: ", text_color="black")
        meals_label.grid(row=2, column=0, padx=(10, 5), pady=(5, 5), sticky="w")
        self.meals_value = CTkLabel(frame, text="", text_color="black")
        self.meals_value.grid(row=2, column=1, padx=(5, 10), pady=(5, 5), sticky="w")

        meals_sub_label = CTkLabel(frame, text="Meal Subscribed To: ", text_color="black")
        meals_sub_label.grid(row=3, column=0, padx=(10, 5), pady=(5, 10), sticky="w")
        self.meals_sub_value = CTkLabel(frame, text="", font=CTkFont(size=16, weight="bold"), text_color="black")
        self.meals_sub_value.grid(row=3, column=1, padx=(5, 10), pady=(5, 10), sticky="w")

        return frame

    def update_details_box(self, name, reg_number, num_meals, meal_list, color):
        self.name_value.configure(text=name)
        self.reg_value.configure(text=reg_number)
        self.meals_value.configure(text=num_meals)
        self.meals_sub_value.configure(text=meal_list)
        self.details_frame.configure(fg_color=color)
        
    def generate_meal_colors(self):
        color_map = {}

        color_map["Breakfast"] = "#ADD8E6"  # Light Blue
        color_map["Lunch"] = "#90EE90"      # Light Green
        color_map["Dinner"] = "#F08080"     # Light Coral

        color_map["Breakfast, Lunch"] = "#E0FFFF"  # Light Cyan
        color_map["Breakfast, Dinner"] = "#FAFAD2" # Light Goldenrod Yellow
        color_map["Lunch, Dinner"] = "#D8BFD8"   # Thistle

        color_map["Breakfast, Lunch, Dinner"] = "#F5F5DC" # Beige

        color_map["LEAVE"] = "#F5A97F"          # Peach
        color_map["NOT"] = "#D3D3D3"            # Light Gray
        color_map["ALREADY_EATEN"] = "#FF6347"   # Tomato
        return color_map

    def create_prepaid_entry(self, batch):
        
        if self.workbook() is None:
            self.write_to_status_bar('No active Daily Entry File found. Please create a new one.')
            return
        
        if batch == 'MS25':
            num = str(self.ms25.get()).rjust(3, '0')
            registration_number = f'MS25{num}'
            self.ms25.delete(0, 'end')
        elif batch == 'MS24':
            num = str(self.ms24.get()).rjust(3, '0')
            registration_number = f'MS24{num}'
            self.ms24.delete(0, 'end')
        elif batch == 'MS23':
            num = str(self.ms23.get()).rjust(3, '0')
            registration_number = f'MS23{num}'
            self.ms23.delete(0, 'end')
        else:
            registration_number = self.others.get().upper().strip()
            self.others.delete(0, 'end')

        offline_prepaid_sheet = self.workbook()['Prepaid Sheet']
        meal_types = ['veg', 'non-veg']

        subscriber_registration_numbers = column_values(offline_prepaid_sheet, 2)
        if registration_number not in subscriber_registration_numbers:
            self.write_to_status_bar(f'{registration_number} has not subscribed to any meal.')
            return

        idx_of_registration_number = subscriber_registration_numbers.index(registration_number) + 2

        subscriber_data = row_values(offline_prepaid_sheet, idx_of_registration_number)
        name = subscriber_data[0]
        
        subscribed_meals = []
        if subscriber_data[2] != "NOT":
            subscribed_meals.append("Breakfast")
        if subscriber_data[4] != "NOT":
            subscribed_meals.append("Lunch")
        if subscriber_data[6] != "NOT":
            subscribed_meals.append("Dinner")

        num_meals_subscribed = len(subscribed_meals)
        meal_string = ", ".join(subscribed_meals)
        
        current_meal_type = MEAL_COLUMN_MAPPING[self.meal.get()]
        status_col = current_meal_type['status']
        time_col = current_meal_type['time']
        current_meal_status = subscriber_data[status_col]
        
        if current_meal_status in meal_types:
            display_color = self.meal_colors["ALREADY_EATEN"]
        elif current_meal_status == 'LEAVE':
            display_color = self.meal_colors["LEAVE"]
        elif current_meal_status == 'NOT':
            display_color = self.meal_colors["NOT"]
        else:
            display_color = self.meal_colors.get(meal_string, "white")

        if self.non_veg.get() == 1:
            current_meal_id = "non-veg"
        else:
            current_meal_id = "veg"

        if current_meal_status in meal_types:
            self.write_to_status_bar(f'{registration_number}: {name} was already checked. STOP!')
            return
        elif current_meal_status == 'LEAVE':
            self.write_to_status_bar(f'{registration_number}: {name} is on LEAVE. STOP!')
            return
        elif current_meal_status == 'NOT':
            self.write_to_status_bar(f'{registration_number}: {name} is not subscribed in this meal. STOP!')
            return
        
        self.update_details_box(name, registration_number, num_meals_subscribed, meal_string, display_color)
        
        current_time = now().strftime("%H:%M:%S")
        
        if self.update.get() == 1 and ONLINE_ENABLED:
            gsheet_prepaid_sheet = self.gsheet().worksheet('Prepaid Sheet')
            online_meal_status = gsheet_prepaid_sheet.cell(idx_of_registration_number, status_col).value
            if online_meal_status in meal_types:
                self.write_to_status_bar(f'{registration_number}: {name} was checked in other mess. STOP!')
                return
            gsheet_prepaid_sheet.update_cell(idx_of_registration_number, status_col + 1, current_meal_id)
            gsheet_prepaid_sheet.update_cell(idx_of_registration_number, time_col + 1, current_time)
            
        offline_prepaid_sheet.cell(idx_of_registration_number, status_col + 1).value = current_meal_id
        offline_prepaid_sheet.cell(idx_of_registration_number, time_col + 1).value = current_time
            
        self.write_to_status_bar(f'{registration_number}: {name} is checked.')
        if self.non_veg.get() == 1:
            self.generate_coupon(name, self.prepaid_extra_price.get())
        else:
            self.workbook().save(self.get_file('daily_entry'))

    def generate_coupon(self, name, price):
        
        if self.workbook() is None:
            self.write_to_status_bar('No active Daily Entry File found. Please create a new one.')
            return
        
        coupon_sheet = self.workbook()[f'Coupons {self.meal.get()}']
        try:
            price_float = float(price)
        except ValueError:
            price_float = 0.0

        details_to_append = [name, price_float, now().strftime("%H:%M:%S")]
        coupon_sheet.append(details_to_append)

        if self.update.get() == 1 and ONLINE_ENABLED:
            coupon_gsheet = self.gsheet().worksheet(f'Coupons {self.meal.get()}')
            coupon_gsheet.append_row(details_to_append)

        self.coupon.delete(0, 'end')
        self.amount.delete(0, 'end')

        self.write_to_status_bar(f'Coupon Generated for {name}.')
        self.coupons_sold.configure(state='normal')
        self.coupons_sold.delete(0, 'end')
        self.coupons_sold.insert(0, coupon_sheet.max_row - 1)
        self.coupons_sold.configure(state='readonly')

        self.workbook().save(self.get_file('daily_entry'))

    def create_daily_file(self):
        
        self.clear_cache()
        
        with open(self.get_file('log'), 'w') as file:
            json.dump([], file)

        if self.update_leave_repository.get() and ONLINE_ENABLED:
            self.write_to_status_bar('Updating Leave Data')
            leave_update()

        if self.update_repository.get() and ONLINE_ENABLED:
            self.write_to_status_bar('Updating Subscriber Data')
            subscriber_data_update()

        subscriber_count = {
            "breakfast": 0,
            "lunch": 0,
            "dinner": 0
        }
        leaves = {
            "breakfast": 0,
            "lunch": 0,
            "dinner": 0
        }

        if not os.path.exists('Subscriber Data.xlsx'):
            self.write_to_status_bar('Subscriber Data File not found!')
            return

        subscriber_data_workbook = xl.load_workbook('Subscriber Data.xlsx')
        subscriber_data_worksheet = subscriber_data_workbook['Subscriber Data']

        if os.path.exists(self.get_file('daily_entry')):
            self.write_to_status_bar('Tomorrow\'s file already exists!')
            return

        student_names = column_values(subscriber_data_worksheet, 1)
        registration_numbers = column_values(subscriber_data_worksheet, 2)

        today_s_workbook = xl.Workbook()
        self._workbook_cache = today_s_workbook
        today_s_workbook.remove(today_s_workbook['Sheet'])
        today_s_workbook.create_sheet('Prepaid Sheet')
        today_s_workbook.create_sheet('Coupons Breakfast')
        today_s_workbook.create_sheet('Coupons Lunch')
        today_s_workbook.create_sheet('Coupons Dinner')
        today_s_workbook.create_sheet('Calculations')
        
        prepaid_sheet = today_s_workbook['Prepaid Sheet']
        coupons_breakfast_sheet = today_s_workbook['Coupons Breakfast']
        coupons_lunch_sheet = today_s_workbook['Coupons Lunch']
        coupons_dinner_sheet = today_s_workbook['Coupons Dinner']

        prepaid_sheet_header = [
            'Student Name', 'Registration Number',
            'Breakfast', 'Breakfast Time',
            'Lunch', 'Lunch Time',
            'Dinner', 'Dinner Time'
        ]
        coupons_sheet_header = ['Registration Number', 'Amount', 'Time']

        prepaid_sheet.append(prepaid_sheet_header)
        coupons_breakfast_sheet.append(coupons_sheet_header)
        coupons_lunch_sheet.append(coupons_sheet_header)
        coupons_dinner_sheet.append(coupons_sheet_header)

        student_details = list(zip(student_names, registration_numbers))

        for idx, (student_name, registration_number) in enumerate(student_details, start=2):
            prepaid_sheet[f'A{idx}'].value = student_name
            prepaid_sheet[f'B{idx}'].value = registration_number.upper().strip()

            breakfast_status = subscriber_data_workbook['Subscriber Data'][f'C{idx}'].value
            if breakfast_status == 'NOT':
                prepaid_sheet[f'C{idx}'].value = 'NOT'
            else:
                subscriber_count['breakfast'] += 1

            lunch_status = subscriber_data_workbook['Subscriber Data'][f'D{idx}'].value
            if lunch_status == 'NOT':
                prepaid_sheet[f'E{idx}'].value = 'NOT'
            else:
                subscriber_count['lunch'] += 1

            dinner_status = subscriber_data_workbook['Subscriber Data'][f'E{idx}'].value
            if dinner_status == 'NOT':
                prepaid_sheet[f'G{idx}'].value = 'NOT'
            else:
                subscriber_count['dinner'] += 1

        if not self.update_leave_repository.get():
            self.write_to_status_bar('Warning! Leave Update is not enabled. Skipping updating leaves')
        else:
            current_leave_details_worksheet = gsheet_client.open('Leave Details for SRM').worksheet('Current Leave Details')
            current_leave_details = current_leave_details_worksheet.get_all_values()
            
            if len(current_leave_details) == 1:
                leaves['breakfast'] = 0
                leaves['lunch'] = 0
                leaves['dinner'] = 0
                self.write_to_status_bar('No leaves found.')
            else:
                current_leave_details = current_leave_details[1:]
                for leave_detail in current_leave_details:
                    registration_number = leave_detail[3].upper().strip()
                    try:
                        idx = registration_numbers.index(registration_number) + 2
                    except ValueError:
                        continue
                    if prepaid_sheet[f'C{idx}'].value != 'NOT':
                        prepaid_sheet[f'C{idx}'].value = 'LEAVE'
                        leaves['breakfast'] += 1
                    if prepaid_sheet[f'E{idx}'].value != 'NOT':
                        prepaid_sheet[f'E{idx}'].value = 'LEAVE'
                        leaves['lunch'] += 1
                    if prepaid_sheet[f'G{idx}'].value != 'NOT':
                        prepaid_sheet[f'G{idx}'].value = 'LEAVE'
                        leaves['dinner'] += 1

        if self.create_database.get() == 1 and ONLINE_ENABLED:
            sheet_name = f'{self.date.get()} {self.file_name.get()}'
            try:
                new_gsheet = gsheet_client.open(sheet_name)
                self._gsheet_cache = new_gsheet
                self.write_to_status_bar('Google Sheet already exists. Using existing sheet.')
            except gspread.exceptions.SpreadsheetNotFound:
                new_gsheet = gsheet_client.create(sheet_name)
                self._gsheet_cache = new_gsheet
                repository_details_worksheet = gsheet_client.open('Repository Details for SRM').worksheet('Sheet1')
                repository = Repository(repository_details_worksheet)
                self.gsheet().share("studentmess@iisermohali.ac.in", perm_type='user', role='writer', notify=False)
                for email in repository.share_to_emails:
                    self.gsheet().share(email, perm_type='user', role='writer', notify=False)
                self.gsheet().add_worksheet('Prepaid Sheet', rows=1000, cols=8)
                self.gsheet().add_worksheet('Coupons Breakfast', rows=1000, cols=3)
                self.gsheet().add_worksheet('Coupons Lunch', rows=1000, cols=3)
                self.gsheet().add_worksheet('Coupons Dinner', rows=1000, cols=3)
                self.gsheet().add_worksheet('Calculations', rows=1000, cols=2)
                self.gsheet().add_worksheet('Log', rows=1000, cols=2)
                self.gsheet().del_worksheet(self.gsheet().sheet1)

                prepaid_gsheet = self.gsheet().worksheet('Prepaid Sheet')
                coupons_breakfast_gsheet = self.gsheet().worksheet('Coupons Breakfast')
                coupons_lunch_gsheet = self.gsheet().worksheet('Coupons Lunch')
                coupons_dinner_gsheet = self.gsheet().worksheet('Coupons Dinner')

                prepaid_gsheet.append_row(prepaid_sheet_header)
                coupons_breakfast_gsheet.append_row(coupons_sheet_header)
                coupons_lunch_gsheet.append_row(coupons_sheet_header)
                coupons_dinner_gsheet.append_row(coupons_sheet_header)

                prepaid_data = []

                prepaid_sheet_rows = list(prepaid_sheet.iter_rows(min_row=2, values_only=True))
                for row in prepaid_sheet_rows:
                    prepaid_data.append(list(row))

                gsheet_batch_upload(prepaid_gsheet, prepaid_sheet_header, prepaid_data)
                self.write_to_status_bar('Google Sheet Created!')

        self.workbook().save(self.get_file('daily_entry'))
        self.write_to_status_bar('File Created!')
        self.information_box.configure(state='normal')
        self.information_box.insert(
            '0.0',
            f"""Subscribers:
• Breakfast: {subscriber_count['breakfast']}
• Lunch: {subscriber_count['lunch']}
• Dinner: {subscriber_count['dinner']}

Leaves:
• Breakfast: {leaves['breakfast']}
• Lunch: {leaves['lunch']}
• Dinner: {leaves['dinner']}

Food to be Prepared:
• Breakfast: {subscriber_count['breakfast'] - leaves['breakfast']}
• Lunch: {subscriber_count['lunch'] - leaves['lunch']}
• Dinner: {subscriber_count['dinner'] - leaves['dinner']}
"""
        )
        self.information_box.configure(state='disabled')

    def calculate(self):
        self.write_to_status_bar('Starting Calculations')
        
        if self.workbook() is None:
            self.write_to_status_bar('No active Daily Entry File found. Please create a new one.')
            return

        def get_meal_info(worksheet, col):
            meal_values = column_values(worksheet, col)[1:]
            meal_statistics = {
                "veg": meal_values.count('veg'),
                "non-veg": meal_values.count('non-veg'),
                "leave": meal_values.count('LEAVE'),
                "not-subscribed": meal_values.count('NOT'),
                "not-availed": len(meal_values) - meal_values.count('veg') - meal_values.count('non-veg') - meal_values.count('LEAVE') - meal_values.count('NOT'),
                "coupon_number": 0,
                "coupon_amount": 0.0
            }
            return meal_statistics

        def process_meal_coupons(worksheet, meal_info):
            coupon_values = column_values(worksheet, 2)[1:]
            meal_info['coupon_number'] = len(coupon_values)
            meal_info['coupon_amount'] = sum(float(coupon) for coupon in coupon_values if coupon)
            return meal_info

        def display_meal_info(meal_info, meal, parent):
            parent += f"""{meal}:
• Veg: {meal_info['veg']}
• Non-Veg: {meal_info['non-veg']}
• Leave: {meal_info['leave']}
• Not Subscribed: {meal_info['not-subscribed']}
• Not Availed: {meal_info['not-availed']}
• Coupons: {meal_info['coupon_number']}
• Coupon Amount: {meal_info['coupon_amount']}
-----x-----\n
"""
            return parent

        is_online_sheet_available = True
        try:
            prepaid_sheet = self.gsheet().worksheet('Prepaid Sheet')
            coupons_breakfast_sheet = self.gsheet().worksheet('Coupons Breakfast')
            coupons_lunch_sheet = self.gsheet().worksheet('Coupons Lunch')
            coupons_dinner_sheet = self.gsheet().worksheet('Coupons Dinner')
            calculations_sheet = self.gsheet().worksheet('Calculations')
        except gspread.exceptions.SpreadsheetNotFound:
            is_online_sheet_available = False
            self.write_to_status_bar('Google Sheet not found, using local file instead.')
            
            if not os.path.exists(self.get_file('daily_entry')):
                self.write_to_status_bar('Daily entry file does not exist. Please create a new file before calculating.')
                return
            
            prepaid_sheet = self.workbook()['Prepaid Sheet']
            coupons_breakfast_sheet = self.workbook()['Coupons Breakfast']
            coupons_lunch_sheet = self.workbook()['Coupons Lunch']
            coupons_dinner_sheet = self.workbook()['Coupons Dinner']
            calculations_sheet = self.workbook()['Calculations']

        breakfast_info = get_meal_info(prepaid_sheet, 3)
        lunch_info = get_meal_info(prepaid_sheet, 5)
        dinner_info = get_meal_info(prepaid_sheet, 7)

        breakfast_info = process_meal_coupons(coupons_breakfast_sheet, breakfast_info)
        lunch_info = process_meal_coupons(coupons_lunch_sheet, lunch_info)
        dinner_info = process_meal_coupons(coupons_dinner_sheet, dinner_info)

        display_str = ""
        display_str = display_meal_info(breakfast_info, "Breakfast", display_str)
        display_str = display_meal_info(lunch_info, "Lunch", display_str)
        display_str = display_meal_info(dinner_info, "Dinner", display_str)

        self.information_box.configure(state='normal')
        self.information_box.delete('0.0', 'end')
        self.information_box.insert('0.0', display_str)
        self.information_box.configure(state='disabled')

        if is_online_sheet_available:
            calculations_sheet.clear()
            calculation_data = [[line] for line in display_str.splitlines()]
            calculations_sheet.update(calculation_data, f'A1:A{len(calculation_data)}')
        else:
            for row in calculations_sheet['A1:H100']:
                for cell in row:
                    cell.value = None
            for i, line in enumerate(display_str.splitlines(), start=1):
                calculations_sheet.acell(row=i, column=1, value=line)
            self.workbook().save(self.get_file('daily_entry'))

        self.write_to_status_bar('Calculations Done!')

    def write_to_status_bar(self, text, level='info'):
        if not os.path.exists(self.get_file('log')):
            with open(self.get_file('log'), 'w') as file:
                json.dump([], file)
        with open(self.get_file('log'), 'r') as file:
            log = json.load(file)

        log.append({
            'time': now().strftime("%H:%M:%S"),
            'message': text,
            'level': level
        })

        with open(self.get_file('log'), 'w') as file:
            json.dump(log, file, indent=2)

        self.status.configure(state='normal')
        self.status.delete(0, 'end')
        self.status.insert(0, text)
        self.status.configure(state='readonly')

        if level == 'error':
            try:
                log_sheet = self.gsheet().worksheet('Log')
                log_sheet.append_row([
                    now().strftime("%H:%M:%S"),
                    text
                ])
            except gspread.exceptions.SpreadsheetNotFound:
                pass

    def get_file(self, classification):
        if classification == 'daily_entry':
            directory = 'Daily Entry'
            file_name = f'{self.date.get()} {self.file_name.get()}.xlsx'
        elif classification == 'log':
            directory = 'Logs'
            file_name = f'{self.date.get()} {self.file_name.get()}.json'
        else:
            raise ValueError('Invalid classification provided.')

        if not os.path.exists(directory):
            os.mkdir(directory)
        return os.path.join(directory, file_name)
    
    def workbook(self):
        if self._workbook_cache is None:
            try:
                self._workbook_cache = xl.load_workbook(self.get_file('daily_entry'))
            except FileNotFoundError:
                self.write_to_status_bar('Local Daily Entry file not found.')
                return None
            except PermissionError:
                self.write_to_status_bar('Permission denied while accessing the file.')
                return None
            except Exception as e:
                self.write_to_status_bar(f'An unexpected error occurred: {e}')
                return None
        return self._workbook_cache
    
    def gsheet(self):
        if self._gsheet_cache is None:
            try:
                sheet_name = f'{self.date.get()} {self.file_name.get()}'
                self._gsheet_cache = gsheet_client.open(sheet_name)
            except gspread.exceptions.SpreadsheetNotFound:
                self.write_to_status_bar('Google Sheet not found!')
                return None
            except NewConnectionError:
                self.write_to_status_bar('Internet does not work, trying to do everything locally')
                return None
            except Exception as e:
                self.write_to_status_bar(f'An unexpected error occurred: {e}')
                return None
        return self._gsheet_cache
    
    def clear_cache(self):
        self._workbook_cache = None
        self._gsheet_cache = None

if __name__ == "__main__":
    app = App()
    app.mainloop()
